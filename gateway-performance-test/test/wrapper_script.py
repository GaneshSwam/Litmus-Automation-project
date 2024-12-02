import requests
import urllib3
urllib3.disable_warnings()
from top_ssh import *
from nats_msg_count_ssh import *
from ssh_loopedge_restart import *
from extract_top_data import *
from extract_nats_msg_data import *
from extract_total_topics import *      
from report_generator_in_excel import *
import time                     #for adding delays
import base64                   #for generating authorization token
import os                       #for setting environment variables 
import csv                      #text to csv
from openpyxl import load_workbook, Workbook #for excel related operations
from tqdm import tqdm           #for providing updates on elapsed time n completed iterations

    # storing values of environment variables 
api_username_token = os.environ.get("LE_API_USERNAME_TOKEN") 
api_password_token = ""
host_ip=os.environ.get("LE_HOST_IP_ADDRESS")
ssh_username = os.environ.get("LE_SSH_USERNAME")
ssh_password = os.environ.get("LE_SSH_PASSWORD")


def generate_basic_auth_token(username, password):              #generates the basic authorization token after taking LE API token as username and pwd values
    credentials = f"{username}:{password}"
    encoded_credentials = base64.b64encode(credentials.encode()).decode()
    auth_token = f"Basic {encoded_credentials}"
    return auth_token

def upload_device_template(num_devices):                #for uploading suitable json device template

    url = f'https://{host_ip}/dm/template/v2?locale=en-US'

    headers = {
        'Content-Type': 'application/json',
        'Authorization': generate_basic_auth_token(api_username_token, api_password_token)
    }

    # Define the size parameter for each device template based on the number of devices
    size_parameters = {1: 398, 5: 1710, 10: 3350, 20: 6650, 50: 16550}

    data = f'{{"size": {size_parameters[num_devices]}}}'

    response = requests.post(url, headers=headers, data=data, verify=False)     #post request for uplooading the template
    #print(response)        #for debugging
    responseData = response.json()
    id_value = 0
    #print("responsData:",responseData)     #for debugging
    if 'id' in responseData:
        id_value = responseData['id']
        #print(f'Stored url ID Value for {num_devices} dev:', id_value)          #display the url extension id

    url = f'https://{host_ip}/dm/template/v2/{id_value}/resume?locale=en-US'       #new url with the extension
    headers = {
        'Upload-Length': '1024199',
        'Upload-Offset': '0',
        'Authorization': generate_basic_auth_token(api_username_token, api_password_token)
    }
    files = {
        'uploadFile': (f'{num_devices}_devices.json', open(f'device templates/{num_devices} devices.json', 'rb'))
    }
 
    response = requests.put(url, headers=headers, files=files, verify=False)        #put request for uploading the template
    if response.status_code == 200:
        print('\n'+f"{num_devices} device template uploaded successfully.")
        res=f"{num_devices} device template uploaded successfully."
    else:
        print('\n'+f"Failed to upload the {num_devices} device template. Status code:", response.status_code)
        res=f"Failed to upload the {num_devices} device template."
    return res

def upload_csv(num_devices, num_tags_list):             #for uploading suitable csv file once a device template is uploaded

    for num_tags in tqdm(num_tags_list):
        time.sleep(10)

        # Upload the device template for the current number of devices
        res=upload_device_template(num_devices)
        #print(res)         #for debugging
        
        time.sleep(60)

        if(res==f"{num_devices} device template uploaded successfully."):
            #for tag csv
            url = f'https://{host_ip}/devicehub/v2'
          
            headers = {
            'Authorization': generate_basic_auth_token(api_username_token, api_password_token)
            }
            operations = '{"operationName":"UploadCsv","variables":{"file":null},"query":"mutation UploadCsv($file: Upload!) {\\n  UploadCsv(file: $file) {\\n    Warnings {\\n      Error\\n      Line\\n      LineNumber\\n    }\\n  }\\n}"}'
            map_data = '{"1":["variables.file"]}'
            
            file_name = f'{num_tags}_tags.csv'
            file_path = rf'csvs/{num_devices}_devices/{file_name}'
            files = {
                '1': (file_name, open(file_path, 'rb'))
            }
            response = requests.post(url, headers=headers, data={'operations': operations, 'map': map_data}, files=files, verify=False)     #post request for uploading tag csv files

            if response.status_code == 200:
                print(f"{file_name} CSV file uploaded successfully.")
                csv_status=f"{file_name} CSV file uploaded successfully."
        
            else:
                print(f"Failed to upload the CSV file {file_name}. Status code:", response.status_code)
                csv_status=f"Failed to upload the CSV file {file_name}."

        # After the response is received, print response content:
            print("Response Content:", response.content)

        #to write the device template number, tags number and expected messages count onto the output text file
        exp_msg_count=num_devices*num_tags+num_devices
        output_file_path=r'output_file.txt'     #first output file is in .txt format, which will be converted to .csv and then finally to .xlsx
        if(num_devices==1 and num_tags==10):
            with open(output_file_path, 'w') as output_file:
                output_file.write(str(num_devices) + ', ' + str(num_tags) + ', ' + str(exp_msg_count)+ ', ')
        else:
            with open(output_file_path, 'a') as output_file:
                output_file.write(str(num_devices) + ', ' + str(num_tags) + ', '+ str(exp_msg_count)+ ', ')

        if(csv_status==f"Failed to upload the CSV file {file_name}." or  response.content==b'{"errors":[{"message":"failed to parse multipart form, request body too large"}],"data":null}'):
            
            with open(output_file_path, 'a') as output_file:
                output_file.write(str("csv upload failed") + ', ' + str("csv upload failed") + ', '+str("csv upload failed") + ', ' + str("csv upload failed") + '\n')
            #perform ssh restart
            commands_to_run = [
            'cd /var/lib/loopedge-dh',
            'rm -r db',
            'timeout 10s systemctl restart loopedge-dh',
            'cd ~'
            ]
        # Combining commands using '&&'
            combined_command = ' && '.join(commands_to_run)

            run_ssh_restart_command(host_ip, ssh_username, ssh_password, combined_command)
            time.sleep(60)
            continue
        
        time.sleep(60)
        
        # Command to get system info
        command_to_run = "top -b -n 3 > top_data.txt"

        run_ssh_top_command(host_ip, ssh_username, ssh_password, command_to_run)

        time.sleep(10)

        # Command for getting messages data on nats
        command_to_run = ' timeout 150s ./nats_tool grub --topic "devicehub.raw.>" > /home/root/msg_data.txt'

        run_ssh_nats_command(host_ip, ssh_username, ssh_password, command_to_run)

        #if(csv_status==f"{file_name} CSV file uploaded successfully." and res==f"{num_devices} device template uploaded successfully."):
        #    extract_number_of_topics(r'msg_data.txt', r'output_file.txt') #get no.of topics, if required

        #for extracting messages count data from the txt file generated from the ssh command
        input_file_path = r'msg_data.txt'
        output_file_path = r'output_file.txt'

        file_contents = read_file(input_file_path)
        max_total_messages_count = find_max_total_messages_count(file_contents)

        if(csv_status==f"{file_name} CSV file uploaded successfully." and res==f"{num_devices} device template uploaded successfully."):
            write_output_to_file(output_file_path, (max_total_messages_count/10))
            print('Messages # Metrics successfully calculated and written to the output file.')
            msg_status='Messages # Metrics successfully calculated and written to the output file.'

        #for extracting % cpu and memory used data from the txt generated from the ssh command
        input_file = r'top_data_latest.txt'
        output_file = r'output_file.txt'

        total_cpu, used_mem, total_mem = extract_metrics_from_file(input_file)

        if(csv_status==f"{file_name} CSV file uploaded successfully." and res==f"{num_devices} device template uploaded successfully." and msg_status=='Messages # Metrics successfully calculated and written to the output file.'):
            if total_cpu is not None and used_mem is not None and total_mem is not None:
                write_metrics_to_file(output_file, total_cpu, used_mem, total_mem)
                print('CPU and memory Metrics successfully calculated and written to the output file.')
                top_status='CPU and memory Metrics successfully calculated and written to the output file.'
            else:
                print('Error: Unable to extract metrics from the input file.')

        #to write the percentages of message count to output txt file
        output_file_path=r'output_file.txt'
        with open(output_file_path, 'a') as output_file:
            output_file.write(str(((max_total_messages_count/(10*exp_msg_count))*100)) + '\n')

        time.sleep(10)

        # Commands to perform restart
        commands_to_run = [
        'cd /var/lib/loopedge-dh',
        'rm -r db',
        'timeout 10s systemctl restart loopedge-dh',
        'cd ~'
        ]
        # Combining commands using '&&'
        combined_command = ' && '.join(commands_to_run)

        if(res==f"{num_devices} device template uploaded successfully." and csv_status==f"{file_name} CSV file uploaded successfully." and msg_status=='Messages # Metrics successfully calculated and written to the output file.' and top_status=='CPU and memory Metrics successfully calculated and written to the output file.'):
            run_ssh_restart_command(host_ip, ssh_username, ssh_password, combined_command)
            time.sleep(60)

# Upload files for different numbers of devices and tags lists
devices_list = [1, 5, 10, 20, 50]
tags_list = [10,50,100,500,1000,5000,10000,50000,100000]

for num_devices in tqdm(devices_list):                    #main loop for controlling no.of devices being uploaded
    upload_csv(num_devices, tags_list)

###           Excel report generation and display part            ###

# Define input and output file paths
input_file = r'output_file.txt'  # Path to the input text file
output_file = r'output_file.csv'  # Path to the output CSV file
# Initialize a list to store the CSV data
csv_data = []

# Read lines from the input text file
with open(input_file, 'r') as text_file:
    lines = text_file.readlines()

# Parse lines to create CSV data with numerical values
for line in lines:
    values = line.strip().split(',')
    row = []

    for value in values:
        try:
            # Attempt to convert each value to a float and add it to the row
            numeric_value = float(value)
            row.append(numeric_value)
        except ValueError:
            # Handle non-numeric values by adding None
            row.append(None)

    csv_data.append(row)

# Write CSV data to an output CSV file
with open(output_file, 'w', newline='') as csv_file:
    csv_writer = csv.writer(csv_file)
    csv_writer.writerows(csv_data)

print(f"Converted {input_file} to {output_file}")

# Read data from CSV file with numerical values
with open('output_file.csv', 'r') as csv_file:
    csv_reader = csv.reader(csv_file)
    data_list = [list(map(lambda x: float(x) if x else None, row)) for row in csv_reader]

# Create a new Excel workbook
output_workbook = Workbook()
output_sheet = output_workbook.active

# Insert data into specific location in the workbook
start_row = 15  # Starting row index
start_column = 2  # Starting column index

for row_data in data_list:
    for column_index, value in enumerate(row_data, start=start_column):
        output_sheet.cell(row=start_row, column=column_index, value=value)
    start_row += 1

# Save the modified workbook as an Excel file
output_workbook.save('output.xlsx')

print("Data from CSV file inserted into 'output.xlsx'")

# Open the existing workbook
workbook = load_workbook('output.xlsx')
sheet = workbook.active

create_initial_sheet(sheet) #function to create intial contents of the excel sheet like gateway information
set_contents_before_chart(sheet)    #to create headers for table and put them in a border

create_line_chart(sheet)        #create line charts

print('Report and charts generated. \n')
# Save the workbook
workbook.save('output.xlsx')