import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

def create_initial_sheet(sheet):
    
    # Create a new workbook and load the active sheet
    workbook = openpyxl.load_workbook('output.xlsx')

    # Merge cells and set content for B2:F2
    sheet.merge_cells('B2:F2')
    sheet['B2'] = 'Gateway Details'

    # Set content for B3 and merge C3:F3
    sheet['B3'] = 'Make'
    sheet.merge_cells('C3:F3')

    # Set content for B4 and merge C4:F4
    sheet['B4'] = 'Models'
    sheet.merge_cells('C4:F4')

    # Set content for B5 and merge C5:F5
    sheet['B5'] = 'Processor'
    sheet.merge_cells('C5:F5')

    # Set content for B6 and merge C6:F6
    sheet['B6'] = 'Memory'
    sheet.merge_cells('C6:F6')

    # Set content for B7 and merge C7:F7
    sheet['B7'] = 'LE version'
    sheet.merge_cells('C7:F7')

    # Apply outside border to the range B2:F7
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for row in range(2, 8):
        for col in range(2, 7):
            cell = sheet[get_column_letter(col) + str(row)]
            cell.border = border_style

    # Center-align the text in the "Gateway Details" cell
    gateway_details_cell = sheet['B2']
    gateway_details_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook
    workbook.save('output.xlsx')

def set_contents_before_chart(sheet):

    # Check if cells B12:H13 are already merged and contain the title
    merged_ranges = sheet.merged_cells.ranges
    if any(cell.coordinate in rng for rng in merged_ranges for cell in rng) and sheet['B12'].value == 'Performance Testing for Data Collection Based on Modbus TCP driver':
        # Do nothing, the cell is already merged and contains the title
        pass
    else:
        # Merge cells and set content for B12:H13
        sheet.merge_cells('B12:H13')
        sheet['B12'] = 'Performance Testing for Data Collection Based on Modbus TCP driver'

    # Set contents for B14:H14 if they are empty
    if sheet['B14'].value is None:
        contents = [
            "# of Devices (MODBUS TCP)",
            "# of Tags per Device",
            "Expected # of Messages/sec",
            "Actual # Messages/sec (avg of 5min)",
            "% CPU",
            "% RAM",
            "% Messages"
        ]

        for col, content in enumerate(contents, start=2):
            cell = sheet[get_column_letter(col) + '14']
            cell.value = content

    # Apply thicker outside border to the range B12:H13
    thicker_border = Border(
        left=Side(border_style="thick"),
        right=Side(border_style="thick"),
        top=Side(border_style="thick"),
        bottom=Side(border_style="thick")
    )

    for row in range(12, 14):
        for col in range(2, 9):
            cell = sheet[get_column_letter(col) + str(row)]
            cell.border = thicker_border

    # Center-align the text in the "Gateway Details" cell if it's not centered
    gateway_details_cell = sheet['B12']
    if gateway_details_cell.alignment.horizontal != 'center':
        gateway_details_cell.alignment = Alignment(horizontal='center', vertical='center')

# for creating line charts
def create_line_chart(sheet):
    axis_title = sheet['C14'].value

    chart_details = [
        {"data_range": (15, 23), "chart_title": "1 Device"},
        {"data_range": (24, 32), "chart_title": "5 Devices"},
        {"data_range": (33, 41), "chart_title": "10 Devices"},
        {"data_range": (42, 50), "chart_title": "20 Devices"},
        {"data_range": (51, 59), "chart_title": "50 Devices"}
    ]

    # Define the number of empty rows between charts
    spacing_rows = 2  # Adjust as needed

    current_row = 1  # Initialize the current row position

    for chart_detail in chart_details:
        start_row, end_row = chart_detail["data_range"]
        chart_title = chart_detail["chart_title"]

        chart = LineChart()

        # Add data to the chart
        for col in range(6, 9):
            data = Reference(sheet, min_col=col, min_row=start_row, max_row=end_row)
            labels = Reference(sheet, min_col=3, min_row=start_row, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)

        # Set chart title and axis titles
        chart.title = chart_title
        chart.x_axis.title = axis_title
        chart.y_axis.title = "Percent"

        # Set legend titles
        chart.legend = None  # Remove default legend
        if sheet['F14'].value:
            chart.series[0].name = sheet['F14'].value  # Legend title for series 1
        if sheet['G14'].value:
            chart.series[1].name = sheet['G14'].value  # Legend title for series 2
        if sheet['H14'].value:
            chart.series[2].name = sheet['H14'].value  # Legend title for series 3

        # Add the chart to the worksheet
        chart_cell = sheet.cell(row=current_row, column=10)  # Adjust the cell as needed
        sheet.add_chart(chart, chart_cell.coordinate)

        # Increment the current row position with spacing
        current_row = end_row + spacing_rows + 2


